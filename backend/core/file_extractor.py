"""
JARVIS File Text Extractor
Extrae texto plano de archivos subidos al bucket para inyectar en el contexto del chat.
Soporta: PDF, DOCX, XLSX/ODS, TXT, código, CSV, JSON, imágenes y más.
Si un archivo es muy pesado o no se puede procesar, devuelve un mensaje claro.
"""
import io
from typing import Optional

from backend.core.storage import download_bytes

MAX_CHARS_PER_FILE = 15_000
MAX_FILE_BYTES = 10 * 1024 * 1024  # 10MB max para extracción


def extract_text_from_file(object_key: str, filename_hint: Optional[str] = None) -> str:
    """Descarga un archivo del bucket/local y extrae su texto plano."""
    data = download_bytes(object_key)
    filename = filename_hint or object_key.split("/")[-1]
    ext = filename.split(".")[-1].lower() if "." in filename else ""

    if len(data) > MAX_FILE_BYTES:
        size_mb = len(data) / (1024 * 1024)
        return (
            f"[⚠️ Archivo demasiado grande para analizar: {filename} ({size_mb:.1f}MB). "
            f"El límite es {MAX_FILE_BYTES / (1024*1024):.0f}MB. "
            f"Considerá dividirlo en archivos más pequeños o pedir un análisis por secciones.]"
        )

    if ext == "pdf":
        return _extract_pdf(data, filename)
    elif ext == "docx":
        return _extract_docx(data, filename)
    elif ext in ("xlsx", "xls", "ods"):
        return _extract_spreadsheet(data, filename, ext)
    elif ext in ("jpg", "jpeg", "png", "gif", "webp"):
        return _extract_image_placeholder(filename)
    elif ext in ("mp3", "wav", "webm", "mp4", "mov"):
        return _extract_media_placeholder(filename, ext)
    elif ext in (
        "txt", "md", "markdown", "log", "csv", "json",
        "py", "js", "ts", "jsx", "tsx", "cpp", "h", "hpp",
        "html", "htm", "css", "scss", "sql",
        "xml", "yaml", "yml", "cfg", "ini", "toml", "env",
    ):
        return _extract_text(data, filename)
    else:
        return _extract_text(data, filename)


def _extract_pdf(data: bytes, filename: str) -> str:
    """Extrae texto de PDF usando pypdf."""
    try:
        from pypdf import PdfReader
    except ImportError:
        return f"[Error: pypdf no instalado para leer PDF: {filename}]"

    try:
        reader = PdfReader(io.BytesIO(data))
        text_parts = []
        total = 0
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
                total += len(page_text)
                if total > MAX_CHARS_PER_FILE:
                    break
        full_text = "\n\n".join(text_parts)
        if len(full_text) > MAX_CHARS_PER_FILE:
            page_count = len(reader.pages)
            full_text = full_text[:MAX_CHARS_PER_FILE] + (
                f"\n\n[... Truncado. PDF de {page_count} páginas con {len(full_text)} caracteres totales. "
                f"Mostrando primeras {MAX_CHARS_PER_FILE}. Pedí un análisis por secciones si necesitás más detalle.]"
            )
        return full_text
    except Exception as e:
        return f"[Error al leer PDF {filename}: {e}]"


def _extract_docx(data: bytes, filename: str) -> str:
    """Extrae texto de Word .docx usando python-docx."""
    try:
        from docx import Document
    except ImportError:
        return f"[Error: python-docx no instalado para leer Word: {filename}. Ejecutá: pip install python-docx]"

    try:
        doc = Document(io.BytesIO(data))
        paragraphs = []
        total = 0
        for p in doc.paragraphs:
            if p.text.strip():
                paragraphs.append(p.text)
                total += len(p.text)
                if total > MAX_CHARS_PER_FILE:
                    paragraphs.append(
                        f"\n[... Truncado. Documento Word con más de {MAX_CHARS_PER_FILE} caracteres. "
                        f"Pedí un análisis por secciones específicas.]"
                    )
                    break
        return "\n".join(paragraphs)
    except Exception as e:
        return f"[Error al leer DOCX {filename}: {e}]"


def _extract_spreadsheet(data: bytes, filename: str, ext: str) -> str:
    """Extrae texto de hojas de cálculo usando openpyxl."""
    try:
        import openpyxl
    except ImportError:
        return f"[Error: openpyxl no instalado para leer hoja de cálculo: {filename}. Ejecutá: pip install openpyxl]"

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = []
        total = 0
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                if row_text.strip():
                    rows.append(row_text)
                    total += len(row_text)
                    if total > MAX_CHARS_PER_FILE:
                        rows.append(f"\n[... Truncado en hoja '{sheet_name}'. Archivo muy extenso.]")
                        break
            sheets.append(f"--- Hoja: {sheet_name} ---\n" + "\n".join(rows))
            if total > MAX_CHARS_PER_FILE:
                break
        wb.close()
        return "\n\n".join(sheets)
    except Exception as e:
        return f"[Error al leer hoja de cálculo {filename}: {e}]"


def _extract_text(data: bytes, filename: str) -> str:
    """Extrae texto de archivos de texto plano."""
    try:
        text = data.decode("utf-8", errors="replace")
        if len(text) > MAX_CHARS_PER_FILE:
            text = text[:MAX_CHARS_PER_FILE] + (
                f"\n\n[... Truncado. El archivo tiene {len(text)} caracteres, "
                f"mostrando primeros {MAX_CHARS_PER_FILE}]"
            )
        return text
    except Exception as e:
        return f"[Error al leer archivo de texto {filename}: {e}]"


def _extract_image_placeholder(filename: str) -> str:
    """Placeholder para OCR de imágenes."""
    return (
        f"[🖼️ Archivo de imagen: {filename}. "
        f"Actualmente no puedo leer texto de imágenes (OCR). "
        f"Si necesitás analizar esta imagen, describime qué contiene y te ayudo.]"
    )


def _extract_media_placeholder(filename: str, ext: str) -> str:
    """Placeholder para archivos de audio/video."""
    type_label = "audio" if ext in ("mp3", "wav", "webm") else "video"
    return (
        f"[🎵 Archivo de {type_label}: {filename}. "
        f"Actualmente no puedo transcribir {type_label} automáticamente. "
        f"Si necesitás analizar su contenido, proporcioname una descripción o transcripción.]"
    )


def build_file_context(file_keys: list[str], filenames: Optional[list[str]] = None) -> str:
    """Construye un bloque de contexto con el contenido de múltiples archivos adjuntos."""
    if not file_keys:
        return ""

    parts = ["=" * 60, "📎 CONTENIDO DE ARCHIVOS ADJUNTOS", "=" * 60]

    for i, key in enumerate(file_keys):
        filename = filenames[i] if filenames and i < len(filenames) else key.split("/")[-1]
        parts.append(f"\n📄 Archivo #{i + 1}: {filename}\n{'─' * 40}")
        try:
            content = extract_text_from_file(key, filename)
            parts.append(content)
        except Exception as e:
            parts.append(f"[Error al procesar {filename}: {e}]")

    parts.append("\n" + "=" * 60)
    parts.append("📨 MENSAJE DEL USUARIO (respondé basándote en los archivos anteriores)")
    parts.append("=" * 60)

    return "\n".join(parts)
